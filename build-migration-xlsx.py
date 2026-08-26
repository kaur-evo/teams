import csv, collections, statistics
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SRC='operators_data_202608141724-noname.csv'; OUT='operator-group-migration.xlsx'
DROP_PREFIX='IT-AGR-'; DROP_FACTORY='IT-Agribios'

raw=[]
with open(SRC,encoding='utf-8-sig') as fh:
    rd=csv.reader(fh,delimiter=';'); next(rd)
    for c in rd:
        if not c or not c[0]: continue
        raw.append({'tenantName':c[0],'operatorId':c[1],'stations':c[4],
                    'factories':c[5],'factoryCount':c[6]})
sp=lambda v:[x.strip() for x in (v or '').split(',') if x.strip()]

recs=[]; yara_touched=0; yara_stations_removed=0; emptied=0; already_empty=0
for r in raw:
    st,fa=sp(r['stations']),sp(r['factories']); note=''
    had_st=bool(st)
    if r['tenantName']=='yara' and DROP_FACTORY in fa:
        n=len([s for s in st if s.startswith(DROP_PREFIX)])
        st=[s for s in st if not s.startswith(DROP_PREFIX)]
        fa=[f for f in fa if f!=DROP_FACTORY]
        yara_touched+=1; yara_stations_removed+=n; note='IT-Agribios stations removed'
    fa=sorted(set(fa))
    if not st:
        group=''; note=(note+'; ' if note else '')+'no stations, no group'
        if had_st: emptied+=1
        else: already_empty+=1
    else: group=', '.join(fa)
    recs.append({'tenantName':r['tenantName'],'operatorId':r['operatorId'],
                 'stations':', '.join(st),'factories':', '.join(fa),
                 'factoryCount':len(fa),'operatorGroup':group,'note':note})

counts=collections.Counter((r['tenantName'],r['operatorGroup']) for r in recs if r['operatorGroup'])
per_tenant=collections.Counter(); [per_tenant.__setitem__(t,per_tenant[t]+1) for t,_ in counts]
gv=sorted(per_tenant.values())
combo=[g for (_,g) in counts if ',' in g]
grouped=[r for r in recs if r['operatorGroup']]

wb=Workbook()
ws=wb.active; ws.title='Operators'
cols=['tenantName','operatorId','stations','factories','factoryCount','operatorGroup','note']
ws.append(cols)
for r in recs: ws.append([r[c] for c in cols])

ws2=wb.create_sheet('Groups'); ws2.append(['tenantName','operatorGroup','operators'])
for (t,g),n in sorted(counts.items(), key=lambda x:(x[0][0],-x[1],x[0][1])): ws2.append([t,g,n])

ws3=wb.create_sheet('Summary'); ws3.append(['Item','Value','Notes'])
S=lambda t: ws3.append([t,'',''])
A=lambda a,b,c='': ws3.append([a,b,c])

S('SOURCE')
A('Source file', SRC)
A('Operator rows in source', len(raw))
A('Tenants in source', len(set(r['tenantName'] for r in raw)))

S('RULE APPLIED')
A('Group name', 'the factories the operator works in, comma-separated, alphabetical',
  'Factory comes from station -> station group -> factory. One group per operator.')
A('Separator', 'comma + space', '122 factory names contain a hyphen, so a hyphen would be unreadable. No factory name contains a comma.')

S('YARA CHANGE')
A('Reason','IT-Agribios has no operators actually on shift','Its stations are removed from Yara operators before grouping.')
A('Operators touched', yara_touched)
A('Station assignments removed', yara_stations_removed, 'Stations IT-AGR-SB1, IT-AGR-SB2, IT-AGR-SB3')
A('Prefix rule check','exact, 0 mismatches','No operator had the factory without an IT-AGR- station, or the reverse.')
A('Yara groups before', 63)
A('Yara groups after', per_tenant.get('yara',0))
A('Operators left with no stations by this change', emptied, 'They had only IT-Agribios stations. No group is created for them.')

S('OPERATORS WITHOUT A GROUP')
A('Total without a group', len(recs)-len(grouped))
A('  had no stations in the source', already_empty, 'Largest: marsmtm 370, cocacolahbc 48, yiotis 41')
A('  emptied by the Yara change', emptied)
A('Handling','left blank, no group created','Column operatorGroup is empty and the note says why. They are not on the Groups sheet.')

S('RESULT')
A('Operators receiving a group', len(grouped))
A('Groups to create', len(counts))
A('Tenants receiving groups', len(per_tenant))
A('  single-factory group names', len(counts)-len(combo))
A('  combination group names', len(combo), 'Operator works across several factories')

S('GROUPS PER TENANT')
A('Average', round(sum(gv)/len(gv),2))
A('Median', int(statistics.median(gv)))
A('Minimum', min(gv))
A('Maximum', max(gv), max(per_tenant, key=per_tenant.get))
A('Tenants with exactly 1 group', sum(1 for v in gv if v==1), f'{round(100*sum(1 for v in gv if v==1)/len(gv))}% of tenants')
A('Tenants with more than 10 groups', sum(1 for v in gv if v>10))
A('Groups holding a single operator', sum(1 for v in counts.values() if v==1))

S('LARGEST TENANTS BY GROUP COUNT')
for t,n in per_tenant.most_common(10): A(f'  {t}', n)

S('NAMING WATCH-OUTS')
A('Tenants whose group is named "Default"', sum(1 for (t,g) in counts if g=='Default'),
  'That is the name of their only factory. Decided: keep it, no special-casing.')
longest=max(counts, key=lambda k: len(k[1]))
A('Longest group name', len(longest[1]), f'{longest[0]}: {longest[1]}')

for sheet,widths in ((ws,[18,12,60,40,12,42,32]),(ws2,[24,46,12]),(ws3,[46,58,86])):
    sheet.freeze_panes='A2'
    for i,w in enumerate(widths,1): sheet.column_dimensions[get_column_letter(i)].width=w
    for c in sheet[1]: c.font=Font(bold=True)
ws.auto_filter.ref=ws.dimensions; ws2.auto_filter.ref=ws2.dimensions
for row in ws3.iter_rows(min_row=2,max_col=1):
    c=row[0]
    if c.value and c.value.isupper(): c.font=Font(bold=True)
wb.save(OUT)
print(f"Operators {len(recs)} | Groups {len(counts)} | Tenants {len(per_tenant)} | ilma grupita {len(recs)-len(grouped)}")
print(f"Yara: {yara_touched} operaatorit, {yara_stations_removed} jaamaseost eemaldatud, {emptied} jäi tühjaks, grupid 63 -> {per_tenant.get('yara',0)}")
