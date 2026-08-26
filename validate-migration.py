"""Independently re-derives the migration from the source and checks the
workbook against it. Written to disagree with the builder, not to mirror it:
every expected value here is recomputed from the CSV rather than imported."""
import csv, collections, sys, re, statistics
from openpyxl import load_workbook

SRC='operators_data_202608141724-noname.csv'
# Validate the file named on the command line, otherwise the newest export.
import glob, os
XL=sys.argv[1] if len(sys.argv)>1 else max(
    glob.glob('operator-group-migration-*.xlsx'), key=os.path.getmtime)
DROP_PREFIX='IT-AGR-'; DROP_FACTORY='IT-Agribios'
fails=[]
def check(name, ok, detail=''):
    print(f"{'PASS' if ok else 'FAIL'}  {name}" + ('' if ok else f"\n      {detail}"))
    if not ok: fails.append(name)

sp=lambda v:[x.strip() for x in (v or '').split(',') if x.strip()]

# ---- rebuild expectation from source ----
src={}
with open(SRC,encoding='utf-8-sig') as fh:
    rd=csv.reader(fh,delimiter=';'); hdr=next(rd)
    for c in rd:
        if not c or not c[0]: continue
        src[(c[0],c[1])]={'stations':sp(c[4]),'factories':sp(c[5]),'factoryCount':c[6]}

exp={}
for (t,oid),v in src.items():
    st,fa=list(v['stations']),list(v['factories'])
    if t=='yara' and DROP_FACTORY in fa:
        st=[s for s in st if not s.startswith(DROP_PREFIX)]
        fa=[f for f in fa if f!=DROP_FACTORY]
    fa=sorted(set(fa))
    exp[(t,oid)]={'stations':st,'factories':fa,'group':', '.join(fa) if st else ''}

print(f'validating: {XL}\n')
wb=load_workbook(XL, read_only=True)
check('Workbook has exactly 2 sheets', wb.sheetnames==['Operators','Groups'], wb.sheetnames)

# ---- sheet 1 ----
ws=wb['Operators']; rows=list(ws.values); head=list(rows[0]); body=rows[1:]
check('Operators header as expected',
      head==['tenantName','operatorId','stations','factories','factoryCount','operatorGroup','note'], head)
check('Operators row count equals source', len(body)==len(src), f'{len(body)} vs {len(src)}')

keys=[(r[0],str(r[1])) for r in body]
check('No duplicate operator rows', len(keys)==len(set(keys)),
      f'{len(keys)-len(set(keys))} duplicates')
check('Every source operator present', set(keys)==set(src.keys()),
      f'missing {len(set(src)-set(keys))}, extra {len(set(keys)-set(src))}')

bad_group=bad_st=bad_fc=bad_blank=0
for r in body:
    k=(r[0],str(r[1])); e=exp.get(k)
    if not e: continue
    if (r[5] or '')!=e['group']: bad_group+=1
    if sp(r[2])!=e['stations']: bad_st+=1
    if int(r[4] or 0)!=len(e['factories']): bad_fc+=1
    if not e['stations'] and (r[5] or ''): bad_blank+=1
check('operatorGroup matches recomputed factories', bad_group==0, f'{bad_group} rows differ')
check('stations match after the Yara removal', bad_st==0, f'{bad_st} rows differ')
check('factoryCount equals number of factories listed', bad_fc==0, f'{bad_fc} rows differ')
check('Operators with no stations have no group', bad_blank==0, f'{bad_blank} rows have a group')

# ---- privacy ----
leak=[c for c in head if re.search(r'name', c, re.I) and c!='tenantName']
check('No operator-name column', not leak, leak)
srcnames=set()
with open(SRC,encoding='utf-8-sig') as fh:
    if 'firstName' in fh.readline(): srcnames.add('firstName')
check('Source itself carries no names', not srcnames, srcnames)

# ---- Yara ----
y=[r for r in body if r[0]=='yara']
check('No IT-AGR- station remains in yara',
      not any(s.startswith(DROP_PREFIX) for r in y for s in sp(r[2])))
check('No IT-Agribios factory remains in yara',
      not any(f==DROP_FACTORY for r in y for f in sp(r[3])))
check('No yara group name mentions IT-Agribios',
      not any(DROP_FACTORY in (r[5] or '') for r in y))
untouched=[k for k in src if k[0]!='yara']
diff=sum(1 for k in untouched
         if sp(dict(zip(head,next(r for r in body if (r[0],str(r[1]))==k)))['stations'])!=src[k]['stations'])
check('Non-yara stations untouched', diff==0, f'{diff} changed')

# ---- sheet 2 ----
ws2=wb['Groups']; g=list(ws2.values); ghead=list(g[0]); gbody=g[1:]
check('Groups header as expected',
      ghead==['tenantName','tenantGroups','operatorGroup','operators'], ghead)
want=collections.Counter((r[0],r[5]) for r in body if r[5])
got={(r[0],r[2]):r[3] for r in gbody}
check('No duplicate tenant+group rows', len(gbody)==len(got), f'{len(gbody)-len(got)} duplicates')
check('Groups sheet lists every distinct group', set(got)==set(want),
      f'missing {len(set(want)-set(got))}, extra {len(set(got)-set(want))}')
check('Operator counts per group are correct',
      all(got.get(k)==v for k,v in want.items()),
      str([k for k,v in want.items() if got.get(k)!=v][:3]))
check('Group operator counts sum to grouped operators',
      sum(got.values())==sum(1 for r in body if r[5]),
      f'{sum(got.values())} vs {sum(1 for r in body if r[5])}')
check('No blank group name on Groups sheet', all(r[2] for r in gbody))

# tenantGroups repeats a per-tenant total, so it has to equal that tenant's
# real number of groups on every row it appears on.
pt_want=collections.Counter(t for t,_ in want)
bad_tg=[(r[0],r[1]) for r in gbody if r[1]!=pt_want[r[0]]]
check('tenantGroups equals the tenant group count on every row',
      not bad_tg, str(bad_tg[:3]))
check('tenantGroups is consistent within each tenant',
      all(len({r[1] for r in gbody if r[0]==t})==1 for t in pt_want))
check('tenantGroups rows per tenant match its stated total',
      all(sum(1 for r in gbody if r[0]==t)==pt_want[t] for t in pt_want))

# ---- figures published to Notion ----
# The summary lives in Notion (Shaping Teams & Operators -> Migration of
# operator groups -> Migration output) rather than in the workbook. Restating
# it here means a change in the data fails this script instead of quietly
# leaving the page wrong.
pt=collections.Counter(t for t,_ in want)
gv=sorted(pt.values())
grouped_n=sum(1 for r in body if r[5])
combo_n=sum(1 for (_,g) in want if ',' in g)
yara_touched=sum(1 for k,v in src.items() if k[0]=='yara' and DROP_FACTORY in v['factories'])
yara_stations=sum(1 for k,v in src.items() if k[0]=='yara'
                  for st in v['stations'] if st.startswith(DROP_PREFIX))
emptied=sum(1 for k,v in src.items() if k[0]=='yara' and v['stations']
            and not [s for s in v['stations'] if not s.startswith(DROP_PREFIX)])
longest=max(want, key=lambda k: len(k[1]))

NOTION={
 'Operator rows in source': (22900, len(src)),
 'Tenants in source': (624, len(set(t for t,_ in src))),
 'Yara operators touched': (300, yara_touched),
 'Yara station assignments removed': (900, yara_stations),
 'Yara groups after': (44, sum(1 for t,_ in want if t=='yara')),
 'Left with no stations by the Yara change': (23, emptied),
 'Operators receiving a group': (22181, grouped_n),
 'Groups to create': (1025, len(want)),
 'Tenants receiving groups': (623, len(pt)),
 'Single-factory group names': (961, len(want)-combo_n),
 'Combination group names': (64, combo_n),
 'Median groups per tenant': (1, int(statistics.median(gv))),
 'Tenants with exactly 1 group': (514, sum(1 for v in gv if v==1)),
 'Tenants with more than 10 groups': (8, sum(1 for v in gv if v>10)),
 'Largest tenant group count': (44, max(gv)),
 'Groups holding a single operator': (100, sum(1 for v in want.values() if v==1)),
 'Total operators skipped': (719, len(body)-grouped_n),
 'Had no stations in the source': (696, sum(1 for v in src.values() if not v['stations'])),
 'Tenants named "Default"': (187, sum(1 for (_,g) in want if g=='Default')),
 'Longest group name length': (99, len(longest[1])),
 # The comparison table at the top of the section on the same page.
 'Table: groups created': (1025, len(want)),
 'Table: operator-group links': (22181, grouped_n),
 'Table: worst tenant yara': (44, pt['yara']),
 'Table: groups holding a single operator': (100, sum(1 for v in want.values() if v==1)),
 'Table: combination group names': (64, combo_n),
 # The Yara narrative, which was wrong twice before and is worth pinning down.
 'Yara: combination groups before': (40, None),
 'Yara: combinations containing IT-Agribios': (40, None),
 'Yara: fallback names already standalone': (18, None),
 'Yara: fallback names that are new': (22, None),
 'Yara: standalone IT-Agribios group operators': (23, None),
}
# The yara figures need the pre-change picture, so they are derived here.
_y=[k for k in src if k[0]=='yara']
_before=collections.Counter(', '.join(sorted(set(src[k]['factories']))) for k in _y)
_solo={g for g in _before if ',' not in g}
_combo=[g for g in _before if ',' in g]
_fallback={', '.join(sorted(set(g.split(', '))-{DROP_FACTORY})) for g in _combo if DROP_FACTORY in g}
NOTION['Yara: combination groups before']=(40, len(_combo))
NOTION['Yara: combinations containing IT-Agribios']=(40, sum(1 for g in _combo if DROP_FACTORY in g))
NOTION['Yara: fallback names already standalone']=(18, len(_fallback & _solo))
NOTION['Yara: fallback names that are new']=(22, len(_fallback - _solo))
NOTION['Yara: standalone IT-Agribios group operators']=(23, _before.get(DROP_FACTORY, 0))

# Claims made in prose on the page that no numeric check covered yet.
allf={f for v in src.values() for f in v['factories']}
NOTION['Prose: factory names containing a hyphen']=(122, sum(1 for f in allf if '-' in f))
NOTION['Prose: factory names containing a comma']=(0, sum(1 for f in allf if ',' in f))
NOTION['Prose: distinct factory names']=(790, len(allf))

top=collections.Counter(t for t,_ in want).most_common(10)
NOTION['Prose: largest tenants list']=(
    'yara 44, bostik 35, jw 18, barrus 15, matrixpack 14, cocacolamaroc 12, '
    'hplush 12, corinth 12, thrace 10, yarabig 9',
    ', '.join(f'{t} {n}' for t,n in top))

nost=collections.Counter(k[0] for k,v in src.items() if not v['stations']).most_common(3)
NOTION['Prose: largest no-station tenants']=(
    'marsmtm 370, cocacolahbc 48, yiotis 41',
    ', '.join(f'{t} {n}' for t,n in nost))
NOTION['Prose: longest group name tenant']=('steelcurtainconsulting', longest[0])

for label,(published,actual) in NOTION.items():
    check(f'Notion figure: {label}', published==actual, f'page says {published}, data says {actual}')
avg=round(sum(gv)/len(gv),2)
check('Notion figure: average groups per tenant', avg==1.65, f'page says 1.65, data says {avg}')

print('\n' + (f'{len(fails)} FAILED: '+', '.join(fails) if fails else 'All checks passed.'))
sys.exit(1 if fails else 0)
